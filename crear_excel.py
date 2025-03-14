import pandas as pd
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt

# Nombre del archivo Excel
archivo_excel = "datos_sensores.xlsx"

def limpiar_y_crear_excel():
    if os.path.exists(archivo_excel):
        os.remove(archivo_excel)  # Elimina el archivo existente

    # Crear un nuevo archivo con encabezados
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Datos Sensores"
    sheet.append(["Hora", "Sensor 1", "Sensor 2", "Sensor 3", "Sensor 4"])  # Encabezados
    workbook.save(archivo_excel)
    workbook.close()

def ajustar_columnas(sheet):
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2  # Espaciado extra
        sheet.column_dimensions[col_letter].width = adjusted_width

def agregar_grafica(sheet):
    datos = pd.read_excel(archivo_excel)

    if len(datos) < 2:
        return  # No graficar si hay pocos datos

    # Seleccionar solo los últimos 30 datos
    datos = datos.tail(30)

    plt.figure(figsize=(6, 3))
    colores = ["red", "blue", "green", "yellow"]
    sensores = ["Sensor 1", "Sensor 2", "Sensor 3", "Sensor 4"]

    for i, sensor in enumerate(sensores, start=1):
        plt.plot(datos.index, datos[sensor], label=f"{sensor}", color=colores[i-1])

    plt.xlabel("Últimos 30 datos")
    plt.ylabel("Valor")
    plt.title("Lectura de sensores con un PLC")
    plt.legend()
    plt.grid(True)

    # Guardar imagen temporalmente
    imagen_nombre = "grafica_sensores.png"
    plt.savefig(imagen_nombre, bbox_inches='tight')
    plt.close()

    # Eliminar todas las imágenes previas en la hoja antes de insertar la nueva
    if sheet._images:
        sheet._images.clear()

    # Calcular la celda donde se insertará la gráfica (dos columnas a la derecha de la última columna ocupada)
    columna_grafica = sheet.max_column + 2
    celda_grafica = sheet.cell(row=2, column=columna_grafica).coordinate

    # Agregar la nueva imagen
    img = Image(imagen_nombre)
    sheet.add_image(img, celda_grafica)

def guardar_datos_excel(valores):
    if not os.path.exists(archivo_excel):
        limpiar_y_crear_excel()

    workbook = load_workbook(archivo_excel)
    sheet = workbook.active

    fila_nueva = sheet.max_row + 1
    datos = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        valores[0], valores[1], valores[2], valores[3]
    ]

    for col, valor in enumerate(datos, start=1):
        sheet.cell(row=fila_nueva, column=col, value=valor)

    ajustar_columnas(sheet)
    agregar_grafica(sheet)

    workbook.save(archivo_excel)
    workbook.close()
